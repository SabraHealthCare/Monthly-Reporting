import pandas as pd
import seaborn as sns
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
import os
import pyodbc
import numpy as np
from calendar import monthrange
import sys
from datetime import datetime, timedelta,date
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import xlrd
import warnings
import streamlit as st
import boto3
from io import BytesIO
from io import StringIO
import base64
from tempfile import NamedTemporaryFile
import time
timestr = time.strftime("%Y%m%d-%H%M%S")

#---------------------------define parameters--------------------------
def get_row_no(dataset,row_header):
    return list(dataset.index).index(row_header)
def get_column_no(dataset,col_header):
    return list(dataset.columns).index(col_header)
def strip_lower_col(series_or_list):
    return(list(map(lambda x: str(x).strip().lower() if x==x else x,series_or_list)))
def strip_upper_col(series_or_list):
    return(list(map(lambda x: str(x).strip().upper() if x==x else x,series_or_list)))

def Read_Account_Mapping(bucket_mapping,mapping_path):
    # read account mapping
    mapping_file =s3.get_object(Bucket=bucket_mapping, Key=mapping_path)
    account_mapping = pd.read_excel(mapping_file['Body'].read(), sheet_name=sheet_name_account_mapping,header=0)
    #convert tenant_account to lower case
    account_mapping["Tenant_account"]=strip_lower_col(account_mapping["Tenant_account"])
    account_mapping["Sabra_second_account"]=strip_upper_col(account_mapping["Sabra_second_account"])
    account_mapping["Sabra_account"]=strip_upper_col(account_mapping["Sabra_account"])
    # remove nan in col Sabra_account
    account_mapping=account_mapping.loc[list(map(lambda x:x==x,account_mapping["Sabra_account"])),\
                                     ["Sabra_account","Tenant_account","Sabra_second_account"]]
    account_mapping=account_mapping.loc[list(map(lambda x:x==x,account_mapping["Tenant_account"])),\
                                     ["Sabra_account","Tenant_account","Sabra_second_account"]]
    account_mapping=account_mapping.drop_duplicates()
    account_mapping=account_mapping.reset_index(drop=True)
    return account_mapping
#-----------------------------------------------------------------------------------------
sheet_name_account_mapping="Account_Mapping"
sheet_name_entity_mapping="Property_Mapping"
sheet_name_BPC_pull="BPC_pull"
sheet_name_format='Format'

s3 = boto3.client('s3')
bucket_mapping="sabramapping"

# drop down list of operator
operatorlist = s3.get_object(Bucket=bucket_mapping, Key="Operator_list.xlsx")
operator_list = pd.read_excel(operatorlist['Body'].read(), sheet_name='Operator_list')

st.title("Sabra HealthCare Reporting App")
st.subheader("Operator name:")
operator= st.selectbox(' ',(operator_list))


if operator!='select operator':
    mapping_path="Mapping/"+operator+"/"+operator+"_Mapping.xlsx"
    BPCpull =s3.get_object(Bucket=bucket_mapping, Key=mapping_path)
    BPC_pull=pd.read_excel(BPCpull['Body'].read(),sheet_name=sheet_name_BPC_pull,header=0)
    BPC_pull=BPC_pull.set_index(["ENTITY","ACCOUNT"])
    account_mapping=Read_Account_Mapping(bucket_mapping,mapping_path)
    entity_mapping_obj =s3.get_object(Bucket=bucket_mapping, Key=mapping_path)
    entity_mapping=pd.read_excel(entity_mapping_obj['Body'].read(),sheet_name=sheet_name_entity_mapping,header=0)
    
    Sabra_detail_accounts_list=['PD_MCR_MGD_CARE','PD_MEDICARE','PD_COMM_INS', 'PD_PRIVATE', 'PD_MEDICAID', 'PD_VETERANS', 'PD_MCA_MGD_CARE', 'PD_OTHER','REV_MCR_MGD_CARE', 'REV_MEDICARE','REV_COMM_INS', 'REV_PRIVATE',
     'REV_MEDICAID', 'REV_VETERANS','REV_MCA_MGD_CARE', 'REV_MEDICARE_B','REV_OTHER', 'T_NURSING','T_DIETARY_RAW', 'T_DIETARY_OTHER','T_HOUSKEEPING', 'T_MAINTENANCE','T_MARKETING', 'T_BAD_DEBT','T_LEGAL', 'T_RE_TAX','T_INSURANCE', 
    'T_GEN_ADMIN_OTHER','T_ANCILLARY_THERAPY', 'T_ANCILLARY_PHARMACY','T_ANCILLARY_OTHER', 'T_EXPENSES','T_MGMT_FEE', 'T_OTHER_OP_EXO','T_DEPR_AMORT', 'T_INT_INC_EXP','T_RENT_EXP', 'T_SL_RENT_ADJ_EXP','T_NURSING_LABOR', 'T_N_CONTRACT_LABOR',
     'T_OTHER_NN_LABOR', 'T_CASH_AND_EQUIV','T_AR_GROSS', 'T_AR_VAL_RES','T_INV', 'T_OTH_CUR_ASSETS','T_TRADE_PAY', 'T_OTHER_CUR_LIAB','T_LOC_OUT', 'T_OTHER_DEBT','T_CAPEX', 'T_AR_WRT_OFF','T_LOC_AVAIL', 'REV_ANCILLARY',
     'REV_CONT_ALLOW', 'T_NURSING_HOURS','T_N_CONTRACT_HOURS', 'T_OTHER_HOURS','G_REV_PRF', 'G_SEQ_SUSPENSION','G_FMAP_FUND', 'G_REV_EXTR_COVID','G_EXP_EXTR_COVID']
    month_dic={1:["January","Jan","01/","1/","-1","-01","/1","/01"],2:["February","Feb","02/","2/","-2","-02","/2","/02"],3:["March","Mar","03/","3/","-3","-03","/3","/03"],4:["April","Apr","04/","4/","-4","-04","/4","/04"],5:["May","05/","5/","-5","-05","/5","/05"],6:["June","Jun","06/","6/","-06","-6","/6","/06"],\
               7:["July","Jul","07/","7/","-7","-07","/7","/07"],8:["August","Aug","08/","8/","-8","-08","/8","/08"],9:["September","Sep","09/","9/","-09","-9","/9","/09"],10:["October","Oct","10/","-10","/10",],11:["November","Nov","11/","-11","/11"],12:["December","Dec","12/","-12","/12"]}
    year_dic={2021:["2021","21"],2022:["2022","22"],2023:["2023","23"],2024:["2024","24"],2025:["2025","25"],2026:["2026","26"],2019:["2019","19"],2018:["2018","18"],2020:["2020","20"]} 
    dropdown_title_account='Map to Sabra Account'
    dropdown_title_entity='Map sheet name to Property'  
    Uploading_date=date.today()
    Uploading_year=Uploading_date.year
    Uploading_Lastyear=Uploading_year-1
    Uploading_month=Uploading_date.month  
#------------------------------------functions------------------------------------
#search tenant account column in P&L
# transfer all the account name(revenue, expense, occ) into lower case
# return col number of tenant account
sheet_name="Delaney_Creek_IS"
def Identify_Tenant_Account_Col(PL,account_mapping,sheet_name):
    for tenantAccount_col_no in range(0,PL.shape[1]):
        #trim and lower case column
        account_column=strip_lower_col(PL.iloc[:,tenantAccount_col_no])
        
        #find out how many tenant accounts match with account_mapping list
        match=[x in  list(account_column) for x in account_mapping["Tenant_account"]]
        #If 50% of accounts match with account_mapping list, identify this col as tenant account.
        if len(match)>0 and sum(x for x in match)/len(match)>0.1:
            return tenantAccount_col_no  
        else:
            # wrong account column,continue search accounts col
            continue
    
    # didn't find accounts col
    print("Can't find account column in sheet—— '"+sheet_name+"'")
def Get_Year(single_string):
    if single_string!=single_string or single_string==None or type(single_string)==float:
        return 0,""
    else:
        for Year in year_dic.keys():
            for Year_keyword in year_dic[Year]:
                if Year_keyword in single_string:
                    return Year,Year_keyword
        return 0,""
def Get_Month_Year(single_string):
    if single_string!=single_string or single_string==None or type(single_string)==float:
        return 0,0
    if type(single_string)==datetime:
        return int(single_string.month),int(single_string.year)
    
    single_string=str(single_string)
    Year,Year_keyword=Get_Year(single_string)
    
    # remove year from string
    single_string=single_string.replace(Year_keyword,"")
 
    for Month in month_dic.keys() :#[01,02,03...12]
        for  Month_keyword in month_dic[Month]: #["Jan","January","01","-1","/1",'1/']
            if Month_keyword.lower() in single_string.lower():
                remaining=single_string.lower().replace(Month_keyword.lower(),"").replace("/","")\
                                .replace("-","").replace(" ","").replace("_","")
                #if there are more than 3 other char in the string, this string maybe not the date 
                if len(remaining)>=3:
                    return 0,0
                else:   
                    return Month,Year
            # only year without month, length>3
            else:
                continue
    return 0,Year      
def Month_continuity_check(month_list):
    inv=[]
    month_list=list(filter(lambda x:x!=0,month_list))
    if len(month_list)==0:
        return False
    else:
        inv=[int(month_list[month_i+1])-int(month_list[month_i]) for month_i in range(len(month_list)-1) ]
        if  len(set(inv))<=2 and all([x in [1,-1,11,-11] for x in set(inv)]):
            #continues month, it is month row
            return True
        else:
            return False
def Year_continuity_check(year_list):
    inv=[]
    year_list=list(filter(lambda x:x!=0,year_list))
    if len(year_list)==0:
        return False
    else:
        inv=[int(year_list[year_i+1])-int(year_list[year_i]) for year_i in range(len(year_list)-1) ]
        if len(set(inv))<=2 and all([x in [1,0,-1] for x in set(inv)]):
            #continues year, it is year row
            return True        
        else:
            return False
# add year to month_header: identify current year/last year giving a list of month
def Add_year_to_header(month_list):
    available_month=list(filter(lambda x:x!=0,month_list))
    
    today=date.today()
    current_year= today.year
    last_year=today.year-1
    if len(available_month)==1:
        
        if datetime.strptime(available_month[0]+"/01/"+current_year,'%m/%d/%Y').date()<today:
            year=current_year
        else:
            year=today.year-1
        return year
     
    year_change=0     
    # month decending  #and available_month[0]<today.month
    if (available_month[0]>available_month[1] and available_month[0]!=12) or \
    (available_month[0]==1 and available_month[1]==12) : 
        date_of_assumption=datetime.strptime(str(available_month[0])+"/01/"+str(current_year),'%m/%d/%Y').date()
        if date_of_assumption<today and date_of_assumption.month<today.month:
            report_year_start=current_year
        elif date_of_assumption>=today:
            report_year_start=last_year
        for i in range(len(available_month)):
            available_month[i]=report_year_start-year_change
            if i<len(available_month)-1 and available_month[i+1]==12:
                year_change+=1
            
    # month ascending   
    elif (available_month[0]<available_month[1] and available_month[0]!=12) \
        or (available_month[0]==12 and available_month[1]==1): #and int(available_month[-1])<today.month
        date_of_assumption=datetime.strptime(str(available_month[-1])+"/01/"+str(current_year),'%m/%d/%Y').date()    
        if date_of_assumption<today:
            report_year_start=current_year
        elif date_of_assumption>=today:
            report_year_start=last_year
        for i in range(-1,len(available_month)*(-1)-1,-1):
   
            available_month[i]=report_year_start-year_change
            if i>len(available_month)*(-1) and available_month[i-1]==12:
                #print("year_change",year_change)
                year_change+=1
    
    else:
        return False
 
    j=0
    for i in range(len(month_list)):
        if month_list[i]!=0:
            month_list[i]=available_month[j]
            j+=1
    return month_list  
# find the Month/year row and return row number
def Identify_Month_Row(PL,tenantAccount_col_no,sheet_name):
 
    PL_row_size=PL.shape[0]
    PL_col_size=PL.shape[1]
    search_row_size=min(20,PL_row_size)
    month_table=pd.DataFrame(0,index=range(search_row_size), columns=range(PL_col_size))
    year_table=pd.DataFrame(0,index=range(search_row_size), columns=range(PL_col_size))
    
    for row_i in range(search_row_size):
        for col_i in range(PL_col_size):
            if type(PL.iloc[row_i,col_i])==float:
                continue
            month_table.iloc[row_i,col_i],year_table.iloc[row_i,col_i]=Get_Month_Year(PL.iloc[row_i,col_i])
            
    year_count=[]
    month_count=[]
    max_len=0
    for row_i in range(search_row_size):
        valid_month=list(filter(lambda x:x!=0,month_table.iloc[row_i,]))
        valid_year=list(filter(lambda x:x!=0,year_table.iloc[row_i,]))
        month_count.append(len(valid_month))
        year_count.append(len(valid_year))
    # didn't find any month in all the rows
    if all(map(lambda x:x==0,month_count)):
        print("Can't identify month/year columns in sheet——'"+sheet_name+"'")   
        return [0],0
    month_sort_index = np.argsort(np.array(month_count))
    year_sort_index = np.argsort(np.array(year_count))
    for month_index_i in range(-1,-4,-1): # only check three of the most possible rows
        
        #month_sort_index[-1] is the index number of month_count in which has max month count
        #month_sort_index[i] is also the index/row number of PL
        
        if month_count[month_sort_index[month_index_i]]>1:
            #check validation of month
            #print(Month_continuity_check(month_table.iloc[month_sort_index[month_index_i],]))
            if Month_continuity_check(month_table.iloc[month_sort_index[month_index_i],]):
                for year_index_i in range(-1,-4,-1):
                    # check validation of year
                    if Year_continuity_check(year_table.iloc[year_sort_index[year_index_i],]) \
                        and year_count[year_sort_index[year_index_i]]==month_count[month_sort_index[month_index_i]]:
                       
                        PL_date_header=year_table.iloc[year_sort_index[year_index_i],].apply(lambda x:str(int(x)))+\
                        month_table.iloc[month_sort_index[month_index_i],].apply(lambda x:"" if x==0 else "0"+str(int(x)) if x<10 else str(int(x)))
                        return PL_date_header,month_sort_index[month_index_i]
                    
                    # all the year rows are not valid, add year to month
                    else:
                        continue
                    # all the year rows are not valid, add year to month
                year_table.iloc[year_sort_index[year_index_i],]=Add_year_to_header(list(month_table.iloc[month_sort_index[month_index_i],]))
                PL_date_header=year_table.iloc[year_sort_index[year_index_i],].apply(lambda x:str(int(x)))+\
                month_table.iloc[month_sort_index[month_index_i],].apply(lambda x:"" if x==0 else "0"+str(int(x)) if x<10 else str(int(x)))
                
                st.write("Fail to identify year in the date header in sheet '"+sheet_name+"'. Filled year as below: ")
                original=PL.iloc[month_sort_index[month_index_i],]
               
                d_str = ''
                for i in range(len(PL_date_header)):
                    if PL_date_header[i]==0 or PL_date_header[i]=="0":
                        continue
                    else:
                        d_str +=str(original[i])+"——"+ str(PL_date_header[i])+"  "
                st.write(d_str)
                return PL_date_header,month_sort_index[month_index_i]
                        
            # month is not continuous, check next one
            else:
                continue
                
        # only one month in header:month and year must exist for one month header
        elif month_count[month_sort_index[month_index_i]]==1:
            # month and year must match 
            st.write("There is only one month in sheet——'"+sheet_name+"'")
            col_month=0
            #find the col number of month
            while(month_table.iloc[month_sort_index[month_index_i],col_month]==0):
                col_month+=1
                
                #if month_table.iloc[month_sort_index[index_i],col_month]!=1:
                #if column of month is smaller than column of account, or there is no year in month, continue 
            if col_month<tenantAccount_col_no or year_table.iloc[month_sort_index[month_index_i],col_month]==0:
                st.write("There is no year in date row in sheet——'"+sheet_name+"'")
                continue
           
            count_num=0
            count_str=0
            for row_month in range(month_sort_index[month_index_i],PL.shape[0]):
                if PL.iloc[row_month,col_month]==None or pd.isna(PL.iloc[row_month,col_month]) or PL.iloc[row_month,col_month]=="":
                    continue
                elif type(PL.iloc[row_month,col_month])==float or type(PL.iloc[row_month,col_month])==int:
                    count_num+=1
                else:
                    count_str+=1
                # count_num/str is count of numous/character data under month
                # for a real month column, numous data is supposed to be more than character data
            if count_str>0 and count_num/count_str<0.8:
                continue
                
            else:
                PL_date_header=year_table.iloc[month_sort_index[month_index_i],].apply(lambda x:str(int(x)))+\
                        month_table.iloc[month_sort_index[month_index_i],].apply(lambda x:"" if x==0 else "0"+str(int(x)) if x<10 else str(int(x)))
                        
                return PL_date_header,month_sort_index[month_index_i]
    st.write("Can't identify date row in P&L for sheet: '"+sheet_name+"'")
    return [0],0
def Upload_file_to_S3(file,bucket,key):
    try:
        s3.upload_fileobj(file,bucket,key)
        st.success('Successfully uploaded to S3')
        return True
    except FileNotFoundError:
        time.sleep(6)
        st.error('Fail to upload to S3')
        return False 
     
def Update_Sheet_inS3(bucket,key,sheet_name,df):  
    mapping_file =s3.get_object(Bucket="sabramapping", Key=key)
    workbook = load_workbook(BytesIO(mapping_file['Body'].read()))
    workbook.remove(workbook[sheet_name_account_mapping])
    new_worksheet = workbook.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        new_worksheet.append(r)
    
    with NamedTemporaryFile() as tmp:
         workbook.save(tmp.name)
         data = BytesIO(tmp.read())
    
    s3.upload_fileobj(data,bucket,key)
    st.success('Successfully uploaded to S3')    
    

def Map_New_Account(PL,account_mapping,sheet_name):
    new_accounts=[x if x not in list(account_mapping["Tenant_account"]) and not x!=x else "" for x in PL.index]
    new_accounts=list(filter(lambda x:x!="",new_accounts))
   
    if len(new_accounts)==0:
        return account_mapping
    maplist=[]
    drop_down_list=["No need to map"]+list(account_mapping["Sabra_account"].unique())
    new_account_len=len(new_accounts)
    for account_i in range(new_account_len):
        maplist.append(st.selectbox(new_accounts[account_i],drop_down_list))
        
    if st.button('Submit account mapping'):
        with st.spinner('Updating mapping...'):
        # update account_mapping list, insert new accounts into account_mapping
            len_mapping=len(account_mapping.index)
            j=0
            for i in range(new_account_len):
                if maplist[i]!="No need to map":
                    account_mapping.loc[len_mapping+j,"Sabra_account"]=maplist[i]
                    account_mapping.loc[len_mapping+j,"Tenant_account"]=new_accounts[i]
                    j+=1
                elif maplist[i]=="No need to map":
                    account_mapping.loc[len_mapping+j,"Sabra_account"]="No need to map"
                    account_mapping.loc[len_mapping+j,"Tenant_account"]=new_accounts[i]
                    j+=1
                   
            # update account_mapping workbook       
            Update_Sheet_inS3(bucket_mapping,mapping_path,sheet_name_account_mapping,account_mapping)
            return account_mapping
            
def Map_New_Entity(BPC_pull,entity_mapping):
    Entity_in_BPC=set(BPC_pull.index.get_level_values('ENTITY'))
    Entity_in_format=list(entity_mapping.loc[entity_mapping["Sheet_Name"]==entity_mapping["Sheet_Name"],'Entity'])
    Missing_Entity=list(filter(lambda x:x not in Entity_in_format,Entity_in_BPC))
    if len(Missing_Entity)==0:
        return entity_mapping
    elif len(Missing_Entity)>0:
        st.write("We couldn't found P&L of below properties, please type the corresponding sheet name of these properties in the right box")
    
    maplist=[]
    for entity_i in range(len(Missing_Entity)):
        maplist.append(st.selectbox(BPC_pull.loc[Missing_Entity[entity_i]]["Property_Name"][0],["No need to map"]+finicial_sheet_list))
   
    # update entity_mapping list: insert new entities into entity_mapping
    if st.button('Submit property mapping'):
        with st.spinner('Updating property mapping...'):
        # update entity_mapping list, insert new entities into entity_mapping
         
            len_entity_mapping=entity_mapping.shape[0]
            j=0
            for i in range(len_mapping):
                if maplist[i]!="No need to map":
                    entity_mapping.loc[len_entity_mapping+j,"Sheet_Name"]=maplist[i]
                    entity_mapping.loc[len_entity_mapping+j,"Entity"]=Missing_Entity[i]                    
                    j+=1
                elif maplist[i]=="No need to map":
                    entity_mapping.loc[len_entity_mapping+j,"Sheet_Name"]="No need to map"
                    entity_mapping.loc[len_entity_mapping+j,"Entity"]=Missing_Entity[i]
                    j+=1
            if j>0:             
            # update account_mapping workbook       
                Update_Sheet_inS3(bucket_mapping,mapping_path,sheet_name_entity_mapping,entity_mapping)
            return entity_mapping


    
def Sheet_Process(sheet_name,account_mapping):
        PL = pd.read_excel(uploaded_file,sheet_name=sheet_name,header=None)
        tenantAccount_col_no=Identify_Tenant_Account_Col(PL,account_mapping,sheet_name)
    
        if tenantAccount_col_no==None:
            st.write("Fail to identify tenant account in sheet '"+sheet_name+"'")
            return False,account_mapping
        date_header=Identify_Month_Row(PL,tenantAccount_col_no,sheet_name)
        if len(date_header[0])==1 and date_header[0]==[0]:
            st.write("didn't find date row")
            return False,account_mapping
        
        PL.columns=date_header[0]
        #tenant_account is index of PL, only keep rows with accounts and columns with valid month
       
        PL=PL.set_index(PL.iloc[:,tenantAccount_col_no].values)
        #remove row above date row and remove column without date col name
        PL=PL.iloc[date_header[1]+1:,PL.columns!='0']
        PL.index=map(lambda x:str(x).lower().strip(),PL.index)
        PL.index.name='Tenant_account'
        
        #if there are duplicated accounts in finicial, only keep the last one
        PL=PL[~PL.index.duplicated(keep='last')]
        #remove rows with nan tenant account
        PL=PL.loc[list(filter(lambda x:x!='nan',PL.index))]
    
        # remove columns what are all zero/blank 
        PL=PL.fillna(0)
        
        PL=PL.loc[:, (PL!= 0).any(axis=0)]
        
        #PL=PL.loc[:,PL.apply(pd.Series.nunique) != 1]
        
       
        #  new accounts don't counted yet
        account_mapping=Map_New_Account(PL,account_mapping,sheet_name)
        
        return PL,account_mapping    
    
def Aggregat_PL(PL,account_mapping,entity):
    # convert index to 0,1,2,3....to avoid duplication, original index:'Tenant_account'
    account_mapping=account_mapping.loc[list(map(lambda x:x!='NO NEED TO MAP',account_mapping["Sabra_account"])),["Sabra_account","Tenant_account","Sabra_second_account"]]
    PL=PL.reset_index(drop=False)
    second_account_mapping=account_mapping[account_mapping["Sabra_second_account"]==account_mapping["Sabra_second_account"]][["Sabra_second_account","Tenant_account"]].\
                            rename(columns={"Sabra_second_account": "Sabra_account"})
    
    PL=pd.concat([PL.merge(second_account_mapping,on='Tenant_account',how='right'),PL.merge(account_mapping[["Sabra_account","Tenant_account"]],on='Tenant_account',how='right')])
    
    PL=PL.set_index('Sabra_account',drop=True)
    
    PL.index.name="Sabra_account"
    PL_with_detail=PL
    # aggregate by sabra_account
    PL=PL.drop('Tenant_account', axis=1)
    PL=PL.groupby(by="Sabra_account").sum()
    
    PL.index=[[entity]*len(PL.index),list(PL.index)]
    PL_with_detail.index=[[entity]*len(PL_with_detail.index),PL_with_detail.index]
    return PL,PL_with_detail

    
    
def Compare_PL_BPC(BPC_pull,Total_PL,entity_mapping,account_mapping):
    diff_BPC_PL=pd.DataFrame(columns=["TIME","Entity","Property_Name","Sabra_account","Sheet_name","Sabra","P&L","Diff"])
    for entity in entity_mapping["Entity"]:
        for matrix in Sabra_detail_accounts_list: 
            for timeid in Total_PL.columns:
                try:
                    BPC_value=int(BPC_pull.loc[entity,matrix][timeid+'00'])
                except:
                    BPC_value=0
                try:
                    Operator_value=int(Total_PL.loc[entity,matrix][timeid])
                except:
                    Operator_value=0
                
                if BPC_value==0 and Operator_value==0:
                    continue
               
                if abs(BPC_value-Operator_value)>3:
                    property_name=entity_mapping.loc[entity_mapping["Entity"]==entity,"Property_Name"].item()
                    sheet_name=entity_mapping.loc[entity_mapping["Entity"]==entity,'Sheet_Name'].item()
                    diff_record=pd.DataFrame({"TIME":timeid,"Entity":entity,"Property_Name":property_name,"Sabra_account":matrix,\
                    "Sheet_name":sheet_name,"Sabra":BPC_value,"P&L":Operator_value,"Diff":BPC_value-Operator_value},index=[0])
                    diff_BPC_PL=pd.concat([diff_BPC_PL,diff_record],ignore_index=True)
    return diff_BPC_PL 

def View_Summary(Total_PL,latest_month):
    
    months=list(Total_PL.columns)
    m_str = ''
    for month in months:
        m_str += " " + month 
    st.write("Reporting months:"+m_str)   
    st.write("The latest reporting month is:"+str(max(months)))
    st.dataframe(Total_PL[str(max(months))])
    download_report(Total_PL[str(max(months))].reset_index(drop=False),operator+" "+str(latest_month)+" Reporting")  

def Diff_plot(diff_BPC_PL,PL_with_detail,Total_PL):   
    num_dismatch=diff_BPC_PL.shape[0]
    num_total_data=Total_PL.shape[0]*Total_PL.shape[1]
    percent_dismatch_accounts=num_dismatch/num_total_data
    st.write("{0:.0f}% P&L data were dispatched with Sabra data".format(percent_dismatch_accounts*100))
    if len(diff_BPC_PL['Property_Name'].unique())==1:
        col1,col2=st.columns(2)
        with col1:
            fig=plt.figure()
            diff_BPC_PL["Sabra_account"].value_counts().plot(kind="bar")
            #plt.xticks(rotation=45)
            st.pyplot(fig)
        with col2:
            fig=plt.figure()
            diff_BPC_PL["TIME"].value_counts().plot(kind="bar")
            st.pyplot(fig)
    else:  
        col1,col2,col3=st.columns(3)
        with col1:
            fig=plt.figure()
            diff_BPC_PL["Property_Name"].value_counts().plot(kind="bar")
            #plt.xticks(rotation=45)
            st.pyplot(fig)
        with col2:
            fig=plt.figure()
            diff_BPC_PL["Sabra_account"].value_counts().plot(kind="bar")
            #plt.xticks(rotation=45)
            st.pyplot(fig)
        with col3:
            fig=plt.figure()
            diff_BPC_PL["TIME"].value_counts().plot(kind="bar")
            st.pyplot(fig)
    
    select_month=st.selectbox("Select Year/Month",diff_BPC_PL['TIME'].unique().tolist())
    select_Sabra_account=st.selectbox("Select Sabra_account",diff_BPC_PL['Sabra_account'].unique().tolist())
    st.write(select_Sabra_account)
    st.write(PL_with_detail[PL_with_detail["Sabra_account"]==select_Sabra_account])
    #selected_data=PL_with_detail[PL_with_detail["Sabra_account"]==select_Sabra_account][select_month]
    #st.dataframe(selected_data)
    download_report(PL_with_detail.reset_index(drop=False),"Detail of dismatch")  
def download_report(df,button_display):
    download_file=df.to_csv(index=False).encode('utf-8')
    st.download_button(label="Download "+button_display,data=download_file,file_name=operator+" "+button_display+".csv",mime="text/csv")

def Upload_Main(entity_mapping,account_mapping):      
        mapping_format =s3.get_object(Bucket=bucket_mapping, Key=mapping_path)
        format_table=pd.read_excel(mapping_format['Body'].read(), sheet_name=sheet_name_format,header=0)

        TENANT_ID=format_table["Tenant_ID"][0]
        Total_PL=pd.DataFrame()
        Total_PL_detail=pd.DataFrame()
        TENANT_ID=format_table["Tenant_ID"][0]
        
        if format_table["Accounts_in_multiple_sheets"][0]=="N" and format_table["Entity_in_multiple_sheets"][0]=="Y":
        #All accounts are in one sheet
        # how about if entity is sold? it is in entity but not in financial anymore
            for entity_i in range(len(entity_mapping['Entity'])):
                sheet_name=str(entity_mapping.loc[entity_i,"Sheet_Name"])
                
                # sheet_name is not nan
                if sheet_name==sheet_name and sheet_name in PL_sheet_list:
                    PL,account_mapping=Sheet_Process(sheet_name,account_mapping)
                    PL,PL_with_detail=Aggregat_PL(PL,account_mapping,entity_mapping.loc[entity_i,"Entity"])
                    Total_PL=pd.concat([Total_PL,PL], ignore_index=False, sort=False)
                    Total_PL_detail=pd.concat([Total_PL_detail,PL_with_detail], ignore_index=False, sort=False)
                    
                elif (sheet_name!=sheet_name or sheet_name not in PL_sheet_list) and entity_i!=len(entity_mapping['Entity'])-1:
                    continue
               
                if entity_i==len(entity_mapping['Entity'])-1:
                    start_date=min(Total_PL.columns)+"00"
                    end_date=max(Total_PL.columns)+"00"
                
                    # if found new entities in BPC which is not in entity_mapping,
                    # ask for mapping and update entity_mapping, re-do sheet process for new entities.
                    entity_mapping=Map_New_Entity(BPC_pull,entity_mapping)
        latest_month=max(list(Total_PL.columns))
        diff_BPC_PL=Compare_PL_BPC(BPC_pull,Total_PL,entity_mapping,account_mapping)
        if diff_BPC_PL.shape[0]==0:
            st.write("100% matches")
            
        else:
            with st.expander("Summary of Checking"):
                View_Summary(Total_PL,latest_month)
            with st.expander("Detail of dismatch"):
                Diff_plot(diff_BPC_PL,PL_with_detail,Total_PL)
            with st.expander("Download Checking Results"):
                col1,col2=st.columns(2)
                with col1:
                    download_report(diff_BPC_PL,"Checking Result")
                with col2:
                    download_report(Total_PL[latest_month],operator+"_{}_"+"Reporting".format(latest_month))  
def Manage_Mapping_Main():
    col1,col2=st.columns(2)
    with col1:
        tenant_account1=st.text_input("Enter new account")
        tenant_account2=st.selectbox("Edit existed account",['']+list(account_mapping["Tenant_account"].unique()))
        tenant_account1=st.text_input("Enter sheetname of new property")
        tenant_account2=st.selectbox("Edit sheetname of existed property",['']+list(entity_mapping["Sheet_Name"].unique()))
    
    
    with col2:   
        Sabra_account1=st.selectbox("Map Sabra account",['']+list(account_mapping["Sabra_account"].unique()))
        Sabra_account2=st.selectbox("Map Sabra account ",['']+list(account_mapping["Sabra_account"].unique()))
        Sabra_account1=st.selectbox("Map property name",['']+list(entity_mapping["Property_Name"].unique()))
        Sabra_account2=st.selectbox("Map property name ",['']+list(entity_mapping["Property_Name"].unique()))
    
        
    if st.button("Submit"):
        with st.expander("New mapping"):
             st.write(new_account)
    with st.expander("View Sabra-{} Property Mapping".format(operator)):
        st.write(entity_mapping)
        download_report(entity_mapping,operator+" Property Mapping")
    with st.expander("View Sabra-{} Accounts Mapping".format(operator)):      
        st.dataframe(account_mapping)
        download_report(account_mapping,operator+" Account Mapping")
#----------------------------------website widges------------------------------------
  
menu=["Upload P&L","Manage Mapping","Instructions"]

choice=st.sidebar.selectbox("Menu",menu)
if choice=="Upload P&L" and operator!='select operator':
    st.subheader("Upload P&L:")
    uploaded_file=st.file_uploader(" ",type={"xlsx", "xlsm","xls"},accept_multiple_files=False)
    if uploaded_file:
        if uploaded_file.name[-5:]=='.xlsx':
            PL_sheet_list=load_workbook(uploaded_file).sheetnames
        
        Upload_Main(entity_mapping,account_mapping)

elif choice=="Manage Mapping" and operator!='select operator':
    st.subheader("Manage Mapping")
    Manage_Mapping_Main()

       
       








